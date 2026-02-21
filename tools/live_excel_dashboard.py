#!/usr/bin/env python3
from __future__ import annotations

import argparse
import datetime as dt
import json
from collections import deque
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover - dependency check
    raise SystemExit("Please install openpyxl first: pip install openpyxl") from exc


HTML = """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>var_gold live charts</title>
  <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
  <style>
    body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; margin: 16px; background: #0b1020; color: #e5e7eb; }
    .meta { margin-bottom: 12px; color: #9ca3af; }
    .card { background: #121a31; border: 1px solid #263353; border-radius: 10px; padding: 12px; margin-bottom: 14px; }
    canvas { width: 100%; height: 320px; }
    a { color: #93c5fd; }
  </style>
</head>
<body>
  <h2>var_gold sheet2 live view</h2>
  <div class="meta">Local only: <code>127.0.0.1</code> | file: <code id="filePath"></code> | rows: <span id="rowCount">-</span> | last refresh: <span id="lastRefresh">-</span></div>
  <div class="card"><canvas id="spreadChart"></canvas></div>
  <div class="card"><canvas id="fundingChart"></canvas></div>

  <script>
    const REFRESH_MS = __REFRESH_MS__;
    const spreadCtx = document.getElementById("spreadChart").getContext("2d");
    const fundingCtx = document.getElementById("fundingChart").getContext("2d");

    const spreadChart = new Chart(spreadCtx, {
      type: "line",
      data: { labels: [], datasets: [
        { label: "spread_open", data: [], borderColor: "#60a5fa", pointRadius: 0, borderWidth: 1.3 },
        { label: "spread_close_abs", data: [], borderColor: "#f59e0b", pointRadius: 0, borderWidth: 1.3 },
        { label: "spread_gap", data: [], borderColor: "#22c55e", pointRadius: 0, borderWidth: 1.3 }
      ] },
      options: { animation: false, parsing: false, scales: { x: { ticks: { color: "#cbd5e1", maxTicksLimit: 8 }, grid: { color: "#334155" } }, y: { ticks: { color: "#cbd5e1" }, grid: { color: "#334155" } } }, plugins: { legend: { labels: { color: "#e2e8f0" } }, title: { display: true, text: "Spread (open/close)", color: "#e2e8f0" } } }
    });

    const fundingChart = new Chart(fundingCtx, {
      type: "line",
      data: { labels: [], datasets: [
        { label: "paxg_funding", data: [], borderColor: "#a78bfa", pointRadius: 0, borderWidth: 1.3 },
        { label: "xaut_funding", data: [], borderColor: "#f87171", pointRadius: 0, borderWidth: 1.3 },
        { label: "funding_diff_raw", data: [], borderColor: "#2dd4bf", pointRadius: 0, borderWidth: 1.3 }
      ] },
      options: { animation: false, parsing: false, scales: { x: { ticks: { color: "#cbd5e1", maxTicksLimit: 8 }, grid: { color: "#334155" } }, y: { ticks: { color: "#cbd5e1" }, grid: { color: "#334155" } } }, plugins: { legend: { labels: { color: "#e2e8f0" } }, title: { display: true, text: "Funding", color: "#e2e8f0" } } }
    });

    function applyData(payload) {
      const labels = payload.labels;
      spreadChart.data.labels = labels;
      fundingChart.data.labels = labels;

      spreadChart.data.datasets[0].data = payload.spread_open;
      spreadChart.data.datasets[1].data = payload.spread_close_abs;
      spreadChart.data.datasets[2].data = payload.spread_gap;

      fundingChart.data.datasets[0].data = payload.paxg_funding;
      fundingChart.data.datasets[1].data = payload.xaut_funding;
      fundingChart.data.datasets[2].data = payload.funding_diff_raw;

      document.getElementById("filePath").textContent = payload.file_path;
      document.getElementById("rowCount").textContent = payload.row_count;
      document.getElementById("lastRefresh").textContent = new Date().toLocaleTimeString();

      spreadChart.update();
      fundingChart.update();
    }

    async function refresh() {
      try {
        const resp = await fetch("/data", { cache: "no-store" });
        if (!resp.ok) throw new Error("HTTP " + resp.status);
        const payload = await resp.json();
        applyData(payload);
      } catch (err) {
        console.error(err);
      }
    }

    refresh();
    setInterval(refresh, REFRESH_MS);
  </script>
</body>
</html>
"""


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _to_label(value: Any) -> str:
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if value is None:
        return ""
    return str(value)


def read_excel_window(xlsx_path: Path, points: int) -> dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb["data"]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    col_index = {name: idx for idx, name in enumerate(headers)}

    required = [
        "timestamp_utc",
        "spread_open",
        "spread_close_abs",
        "spread_gap",
        "paxg_funding",
        "xaut_funding",
        "funding_diff_raw",
    ]
    missing = [k for k in required if k not in col_index]
    if missing:
        raise ValueError(f"Missing required columns in sheet 'data': {', '.join(missing)}")

    labels = deque(maxlen=points)
    spread_open = deque(maxlen=points)
    spread_close_abs = deque(maxlen=points)
    spread_gap = deque(maxlen=points)
    paxg_funding = deque(maxlen=points)
    xaut_funding = deque(maxlen=points)
    funding_diff_raw = deque(maxlen=points)
    row_count = 0

    for row in rows:
        row_count += 1
        labels.append(_to_label(row[col_index["timestamp_utc"]]))
        spread_open.append(_to_float(row[col_index["spread_open"]]))
        spread_close_abs.append(_to_float(row[col_index["spread_close_abs"]]))
        spread_gap.append(_to_float(row[col_index["spread_gap"]]))
        paxg_funding.append(_to_float(row[col_index["paxg_funding"]]))
        xaut_funding.append(_to_float(row[col_index["xaut_funding"]]))
        funding_diff_raw.append(_to_float(row[col_index["funding_diff_raw"]]))

    wb.close()
    return {
        "file_path": str(xlsx_path),
        "row_count": row_count,
        "labels": list(labels),
        "spread_open": list(spread_open),
        "spread_close_abs": list(spread_close_abs),
        "spread_gap": list(spread_gap),
        "paxg_funding": list(paxg_funding),
        "xaut_funding": list(xaut_funding),
        "funding_diff_raw": list(funding_diff_raw),
    }


def make_handler(xlsx_path: Path, points: int, refresh_ms: int):
    class Handler(BaseHTTPRequestHandler):
        def do_GET(self) -> None:  # noqa: N802
            parsed = urlparse(self.path)
            if parsed.path == "/":
                body = HTML.replace("__REFRESH_MS__", str(refresh_ms)).encode("utf-8")
                self._send(HTTPStatus.OK, "text/html; charset=utf-8", body)
                return
            if parsed.path == "/data":
                try:
                    payload = read_excel_window(xlsx_path, points)
                    body = json.dumps(payload, ensure_ascii=True).encode("utf-8")
                    self._send(HTTPStatus.OK, "application/json; charset=utf-8", body)
                except Exception as exc:  # pragma: no cover - runtime path
                    err = {"error": str(exc)}
                    body = json.dumps(err, ensure_ascii=True).encode("utf-8")
                    self._send(HTTPStatus.INTERNAL_SERVER_ERROR, "application/json; charset=utf-8", body)
                return
            self._send(HTTPStatus.NOT_FOUND, "text/plain; charset=utf-8", b"Not found")

        def _send(self, status: HTTPStatus, content_type: str, body: bytes) -> None:
            self.send_response(status.value)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(len(body)))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(body)

        def log_message(self, fmt: str, *args: Any) -> None:
            # Keep terminal output clean; Ctrl+C still stops the server.
            return

    return Handler


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Serve live web charts from var_gold Excel (sheet2 chart data)."
    )
    parser.add_argument(
        "--xlsx",
        default="/Users/molly/Downloads/var_gold_2026-02-18.xlsx",
        help="Path to the Excel file.",
    )
    parser.add_argument("--host", default="127.0.0.1", help="Bind host. Keep 127.0.0.1 for private local view.")
    parser.add_argument("--port", type=int, default=8787, help="Bind port.")
    parser.add_argument("--points", type=int, default=1200, help="Only render latest N points for performance.")
    parser.add_argument("--refresh-sec", type=float, default=2.0, help="Browser auto-refresh interval.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    xlsx_path = Path(args.xlsx).expanduser().resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")
    handler = make_handler(xlsx_path=xlsx_path, points=max(50, args.points), refresh_ms=max(300, int(args.refresh_sec * 1000)))
    server = ThreadingHTTPServer((args.host, args.port), handler)
    print(f"Serving: http://{args.host}:{args.port}")
    print(f"Excel : {xlsx_path}")
    print("Press Ctrl+C to stop.")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
